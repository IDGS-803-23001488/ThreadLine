import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


def exportar_excel(nombre_archivo, nombre_hoja, headers, data):

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    # ==============================
    # HEADERS
    # ==============================
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alignment = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # ==============================
    # FILAS
    # ==============================
    for row in data:
        ws.append(row)

    # ==============================
    # BORDES
    # ==============================
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border

    # ==============================
    # AUTO WIDTH
    # ==============================
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    # ==============================
    # TABLA ESTILO EXCEL
    # ==============================
    last_col = ws.cell(row=1, column=len(headers)).column_letter

    tab = Table(
        displayName="TablaDatos",
        ref=f"A1:{last_col}{ws.max_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)

    # ==============================
    # OUTPUT
    # ==============================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name=f"{nombre_archivo}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )